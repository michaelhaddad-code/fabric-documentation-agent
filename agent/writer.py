"""
Write all collected data into the Fabric doc template, preserving its formatting.

Tab layout (preserving original + additions):
  Read Me           — untouched
  Lineage           — two blocks (Gold→Silver | Silver→Bronze) + Notes column
  Fabric Artifacts  — + Last Modified, Created By  (cols E, F)
  Tables            — + Last Updated, Source Artifact  (cols F, G)
  Columns           — + Is Nullable  (col G)
  Data Sources      — new tab (Source Name, Type, Connection/Path, Target Bronze, Ingestion Method, Notes)
"""
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    ArtifactInfo, LakehouseInfo, TableInfo, ColumnInfo,
    LineageEdge, DataSourceRow,
)

# Template header colours (matching the dark-blue theme)
_HEADER_FILL = PatternFill('solid', fgColor='2E5A88')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_TITLE_FONT  = Font(color='2E5A88', bold=True)   # block titles on Lineage tab


def _h(ws, row: int, col: int, value: str):
    """Write a header cell with the standard blue styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = copy(_HEADER_FILL)
    cell.font = copy(_HEADER_FONT)
    return cell


def _t(ws, row: int, col: int, value: str):
    """Write a block-title cell (blue text, no fill) — used on Lineage tab."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = copy(_TITLE_FONT)
    return cell


def _sanitize(value):
    """Strip Excel-illegal control characters from string values."""
    if not isinstance(value, str):
        return value
    return ''.join(c for c in value if ord(c) >= 32 or c in '\t\n\r')


def _v(ws, row: int, col: int, value):
    """Write a plain data cell."""
    cell = ws.cell(row=row, column=col, value=_sanitize(value))
    cell.font = Font()  # reset any inherited italic from example rows
    return cell


def _clear_rows(ws, from_row: int):
    """Clear content from from_row to the current max_row."""
    for row in ws.iter_rows(min_row=from_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()


# ---------------------------------------------------------------------------
# Artifacts tab
# ---------------------------------------------------------------------------
# Cols: A=Name, B=Type, C=Workspace, D=Description

def write_artifacts(ws, artifacts: list[ArtifactInfo]):
    _clear_rows(ws, from_row=3)

    _h(ws, 2, 5, 'Last Modified')
    _h(ws, 2, 6, 'Created By')

    for i, art in enumerate(artifacts, start=3):
        _v(ws, i, 1, art.display_name)
        _v(ws, i, 2, art.type)
        _v(ws, i, 3, art.workspace_name)
        _v(ws, i, 4, art.description)
        _v(ws, i, 5, art.last_modified)
        _v(ws, i, 6, art.created_by)


# ---------------------------------------------------------------------------
# Tables tab
# ---------------------------------------------------------------------------
# Original cols: A=Table Name, B=LH/WH Name, C=Layer, D=Description, E=Row Count
# Added:         F=Last Updated, G=Source Artifact

def write_tables(ws, tables: list[TableInfo]):
    _clear_rows(ws, from_row=3)

    _h(ws, 2, 6, 'Last Updated')
    _h(ws, 2, 7, 'Source Artifact')

    for i, tbl in enumerate(tables, start=3):
        _v(ws, i, 1, tbl.table_name)
        _v(ws, i, 2, tbl.lh_name)
        _v(ws, i, 3, tbl.layer)
        _v(ws, i, 4, tbl.description)
        _v(ws, i, 5, tbl.row_count)
        _v(ws, i, 6, tbl.last_updated)
        _v(ws, i, 7, tbl.source_artifact)


# ---------------------------------------------------------------------------
# Columns tab
# ---------------------------------------------------------------------------
# Original cols: A=LH/WH Name, B=Table Name, C=Column Name, D=Datatype, E=Sample Value, F=% Null
# Added:         G=Is Nullable

def write_columns(ws, columns: list[ColumnInfo]):
    _clear_rows(ws, from_row=3)

    _h(ws, 2, 7, 'Is Nullable')

    for i, col in enumerate(columns, start=3):
        _v(ws, i, 1, col.lh_name)
        _v(ws, i, 2, col.table_name)
        _v(ws, i, 3, col.col_name)
        _v(ws, i, 4, col.datatype)
        _v(ws, i, 5, col.sample_value)
        _v(ws, i, 6, col.pct_null)
        _v(ws, i, 7, 'Yes' if col.is_nullable else 'No')


# ---------------------------------------------------------------------------
# Lineage tab
# ---------------------------------------------------------------------------
# Original layout:
#   Row 1: block titles (A1="GOLD -> SILVER", F1="SILVER -> BRONZE")
#   Row 2: headers A-D (left block), F-I (right block), E=gap
#   Row 3+: data
# Added: J=Lineage Notes (covers both blocks)
#
# Left block  (Gold→Silver):  A=Gold LH/WH  B=Gold Table  C=Notebook  D=Silver Table(s)
# Right block (Silver→Bronze): F=Silver LH/WH  G=Silver Table  H=Notebook  I=Bronze Source(s)

def write_lineage(
    ws,
    gold_silver: list[LineageEdge],
    silver_bronze: list[LineageEdge],
):
    _clear_rows(ws, from_row=3)

    # Block titles (row 1)
    _t(ws, 1, 1, 'GOLD -> SILVER')
    _t(ws, 1, 6, 'SILVER -> BRONZE')

    # Headers (row 2) — re-write to ensure consistency + add Notes
    for col, label in [
        (1, 'Gold Layer Lakehouse/Warehouse'),
        (2, 'Gold Layer Table'),
        (3, 'Data Flow / Notebook'),
        (4, 'Silver Layer Table(s)'),
        (6, 'Silver Layer Lakehouse/Warehouse'),
        (7, 'Silver Layer Table'),
        (8, 'Data Flow / Notebook'),
        (9, 'Bronze Layer Table(s)'),
        (10, 'Lineage Notes'),
    ]:
        _h(ws, 2, col, label)

    # The two blocks have independent row counts — zip and fill blanks
    max_rows = max(len(gold_silver), len(silver_bronze))

    for i in range(max_rows):
        row = i + 3

        if i < len(gold_silver):
            e = gold_silver[i]
            _v(ws, row, 1, e.output_lh)
            _v(ws, row, 2, e.output_table)
            _v(ws, row, 3, e.notebook_name)
            _v(ws, row, 4, ', '.join(e.input_refs))

        if i < len(silver_bronze):
            e = silver_bronze[i]
            _v(ws, row, 6, e.output_lh)
            _v(ws, row, 7, e.output_table)
            _v(ws, row, 8, e.notebook_name)
            _v(ws, row, 9, ', '.join(e.input_refs))

        # Notes column: collect from whichever block(s) have a row here
        notes_parts = []
        if i < len(gold_silver) and gold_silver[i].notes:
            notes_parts.append(f'G→S: {gold_silver[i].confidence} — {gold_silver[i].notes}')
        if i < len(silver_bronze) and silver_bronze[i].notes:
            notes_parts.append(f'S→B: {silver_bronze[i].confidence} — {silver_bronze[i].notes}')
        if notes_parts:
            _v(ws, row, 10, '\n'.join(notes_parts))


# ---------------------------------------------------------------------------
# Data Sources tab (new)
# ---------------------------------------------------------------------------
# Cols: A=Source Name, B=Type, C=Connection/Path, D=Target Bronze Table/File,
#       E=Ingestion Method, F=Notes

def add_data_sources_tab(wb, data_sources: list[DataSourceRow]):
    # Create after Columns tab
    ws = wb.create_sheet('Data Sources')

    # Row 1 blank (matches other tabs)
    # Row 2 headers
    for col, label in [
        (1, 'Source Name'),
        (2, 'Type'),
        (3, 'Connection / Path'),
        (4, 'Target Bronze Table / File'),
        (5, 'Ingestion Method'),
        (6, 'Notes'),
    ]:
        _h(ws, 2, col, label)

    for i, src in enumerate(data_sources, start=3):
        _v(ws, i, 1, src.source_name)
        _v(ws, i, 2, src.source_type)
        _v(ws, i, 3, src.connection_path)
        _v(ws, i, 4, src.target_bronze)
        _v(ws, i, 5, src.ingestion_method)
        _v(ws, i, 6, src.notes)

    # Auto-size columns (best effort)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def write_output(
    template_path: Path,
    output_path: Path,
    artifacts: list[ArtifactInfo],
    tables: list[TableInfo],
    columns: list[ColumnInfo],
    gold_silver: list[LineageEdge],
    silver_bronze: list[LineageEdge],
    data_sources: list[DataSourceRow],
):
    wb = openpyxl.load_workbook(template_path)

    write_lineage(wb['Lineage'], gold_silver, silver_bronze)
    write_artifacts(wb['Fabric Artifacts'], artifacts)
    write_tables(wb['Tables'], tables)
    write_columns(wb['Columns'], columns)
    add_data_sources_tab(wb, data_sources)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'\nSaved: {output_path}')
